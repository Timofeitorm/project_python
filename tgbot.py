# --- подключение либ ---
import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.filters import Command
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import aiosqlite
from datetime import datetime, timedelta
import logging
import re
import openpyxl
from io import BytesIO
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN = '7844365314:AAHLa4M51ocIC0EPFqKPCKu8Pln6caGgq74'
bot = Bot(token=TOKEN)
dp = Dispatcher()


# --- Форматирование чисел с пробелами ---
def format_number(num):
    return "{:,.0f}".format(num).replace(",", " ")


# --- Состояния FSM (Finite State Machine — конечный автомат для управления состоянием пользователя)---
class TransactionStates(StatesGroup):
    waiting_for_category = State() #бот ждёт, что пользователь введёт категорию транзакции
    waiting_for_amount = State() #бот ожидает сумму транзакции
    waiting_for_date_range = State()#бот ждёт, выберет ли пользователь диапазон дат
    waiting_for_start_date = State() #если пользователь хочет задать диапазон вручную, бот ждёт начальную дату
    waiting_for_end_date = State() #после ввода начальной даты бот ждёт конечную дату
    waiting_for_category_filter = State() #если бот фильтрует транзакции, он может ждать выбора категории для фильтрации


# --- Инициализация базы данных ---
async def init_db():
    async with aiosqlite.connect("finance.db") as db:
        # Создаем таблицу транзакций (если не существует)
        await db.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                type TEXT,
                amount REAL,
                category TEXT,
                description TEXT DEFAULT '',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Проверяем и добавляем столбец description, если его нет
        try:
            await db.execute('ALTER TABLE transactions ADD COLUMN description TEXT DEFAULT ""')
            await db.commit()
        except aiosqlite.OperationalError as e:
            if "duplicate column name" not in str(e):
                raise e

        # Таблица категорий
        await db.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                user_id INTEGER,
                category_type TEXT,
                category_name TEXT,
                PRIMARY KEY (user_id, category_type, category_name)
            )
        ''')

        # Индексы
        await db.execute('''
            CREATE INDEX IF NOT EXISTS idx_transactions_user_id ON transactions (user_id)
        ''')
        await db.execute('''
            CREATE INDEX IF NOT EXISTS idx_transactions_timestamp ON transactions (timestamp)
        ''')

        await db.commit()
# --- Кнопки ---
def get_main_menu_buttons():#начальное меню
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💸 Добавить расход", callback_data="add_expense"),
         InlineKeyboardButton(text="💰 Добавить доход", callback_data="add_income")],
        [InlineKeyboardButton(text="📊 Отчёты", callback_data="open_reports"),
         InlineKeyboardButton(text="📁 Excel отчёт", callback_data="excel_report")],
        [InlineKeyboardButton(text="🗑 Удалить транзакцию", callback_data="delete_transaction")]
    ])


def get_report_period_buttons():#отчеты
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📅 День", callback_data="r_day"),
         InlineKeyboardButton(text="🗓 Месяц", callback_data="r_month")],
        [InlineKeyboardButton(text="📊 Квартал", callback_data="r_quarter"),
         InlineKeyboardButton(text="📆 Год", callback_data="r_year")],
        [InlineKeyboardButton(text="📅 Ручной период", callback_data="r_custom")],
        [InlineKeyboardButton(text="🔙 Главное меню", callback_data="back_to_menu")]
    ])


def get_filter_buttons(period):#фильтры отчетов
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📈 Только доходы", callback_data=f"f_income_{period}"),
         InlineKeyboardButton(text="📉 Только расходы", callback_data=f"f_expense_{period}")],
        [InlineKeyboardButton(text="📊 Всё", callback_data=f"f_all_{period}")],
        [InlineKeyboardButton(text="🏷 По категории", callback_data=f"f_category_{period}")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="open_reports")]
    ])


def get_cancel_button():#отмена
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_input")]
    ])


def get_back_to_menu_button():#кнопка глаавного меню
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔙 Главное меню", callback_data="back_to_menu")]
    ])


def get_confirmation_buttons(action):#кнопки подтверждения удаления данных
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Да", callback_data=f"confirm_{action}"),
         InlineKeyboardButton(text="❌ Нет", callback_data="cancel_input")]
    ])


# --- Старт ---
@dp.message(Command("start"))
async def start(message: types.Message):
    await message.answer( #начальное сообщение пользователю
        "Привет! Я бот для учёта доходов и расходов.\n"
        "Используй кнопки ниже для управления:",
        reply_markup=get_main_menu_buttons()
    )


# --- Обработка нажатий ---
@dp.callback_query()
async def handle_callback(callback: types.CallbackQuery, state: FSMContext):
    try:
        await callback.answer()
        data = callback.data
        user_id = callback.from_user.id

        if data == "back_to_menu":
            await state.clear()
            await callback.message.answer("Главное меню:", reply_markup=get_main_menu_buttons())

        elif data == "cancel_input":
            await state.clear()
            await callback.message.answer("Действие отменено.", reply_markup=get_main_menu_buttons())

        elif data == "add_expense":
            await state.set_state(TransactionStates.waiting_for_category)
            await state.update_data(t_type="расход")
            await callback.message.answer(
                "Введите категорию расхода (например: еда, транспорт, развлечения):",
                reply_markup=get_cancel_button()
            )

        elif data == "add_income":
            await state.set_state(TransactionStates.waiting_for_category)
            await state.update_data(t_type="доход")
            await callback.message.answer(
                "Введите категорию дохода (например: зарплата, подарок, премия):",
                reply_markup=get_cancel_button()
            )

        elif data == "open_reports":
            await callback.message.answer("Выбери период:", reply_markup=get_report_period_buttons())

        elif data == "excel_report":
            await state.set_state(TransactionStates.waiting_for_start_date)
            await state.update_data(excel_report=True)
            await callback.message.answer(
                "Введите начальную дату для отчета в формате ГГГГ-ММ-ДД:",
                reply_markup=get_cancel_button()
            )

        elif data.startswith("r_"):
            period = data[2:]
            if period == "custom":
                await state.set_state(TransactionStates.waiting_for_start_date)
                await callback.message.answer(
                    "Введите начальную дату в формате ГГГГ-ММ-ДД:",
                    reply_markup=get_cancel_button()
                )
            else:
                await callback.message.answer("Выбери фильтр:", reply_markup=get_filter_buttons(period))

        elif data.startswith("f_"):
            parts = data.split("_")
            f_type = parts[1]  # income / expense / all / category
            period = parts[2]

            if f_type == "category":
                await state.set_state(TransactionStates.waiting_for_category_filter)
                await state.update_data(report_period=period, filter_type=f_type)

                # Получаем список категорий для пользователя
                async with aiosqlite.connect("finance.db") as db:
                    cursor = await db.execute("""
                        SELECT DISTINCT category FROM transactions 
                        WHERE user_id = ? ORDER BY category
                    """, (user_id,))
                    categories = await cursor.fetchall()

                if categories:
                    buttons = []
                    for category in categories:
                        buttons.append([InlineKeyboardButton(
                            text=category[0],
                            callback_data=f"set_category_{period}_{category[0]}"
                        )])
                    buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data=f"open_reports")])

                    await callback.message.answer(
                        "Выберите категорию:",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
                    )
                else:
                    await callback.message.answer(
                        "У вас нет сохраненных категорий.",
                        reply_markup=get_main_menu_buttons()
                    )
            else:
                await send_report(callback, period, f_type)

        elif data.startswith("set_category_"):
            parts = data.split("_")
            period = parts[2]
            category = "_".join(parts[3:])
            await send_report(callback, period, "category", category)

        elif data == "delete_transaction":
            await callback.message.answer(
                "Что вы хотите удалить?",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="💰 Доход", callback_data="del_type_доход")],
                    [InlineKeyboardButton(text="💸 Расход", callback_data="del_type_расход")],
                    [InlineKeyboardButton(text="🔙 Главное меню", callback_data="back_to_menu")]
                ])
            )

        elif data.startswith("del_type_"):
            t_type = data.split("_")[2]  # доход / расход

            async with aiosqlite.connect("finance.db") as db:
                cursor = await db.execute("""
                    SELECT id, amount, category, timestamp
                    FROM transactions
                    WHERE user_id = ? AND type = ?
                    ORDER BY timestamp DESC
                    LIMIT 10
                """, (user_id, t_type))
                rows = await cursor.fetchall()

            if not rows:
                await callback.message.answer(f"📭 Нет записей для удаления ({t_type})",
                                              reply_markup=get_main_menu_buttons())
                return

            buttons = []
            for row in rows:
                tid, amount, category, timestamp = row
                label = f"{format_number(amount)}₽ — {category} ({timestamp[:10]})"
                buttons.append([InlineKeyboardButton(text=label, callback_data=f"del_confirm_{tid}")])

            buttons.append([InlineKeyboardButton(text="🔙 Назад", callback_data="delete_transaction")])

            await callback.message.answer(
                f"Выберите, что удалить ({t_type}):",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
            )

        elif data.startswith("del_confirm_"):
            tid = int(data.split("_")[2])
            await callback.message.answer(
                "Вы уверены, что хотите удалить эту транзакцию?",
                reply_markup=get_confirmation_buttons(f"delete_{tid}")
            )

        elif data.startswith("confirm_delete_"):
            tid = int(data.split("_")[2])
            async with aiosqlite.connect("finance.db") as db:
                await db.execute("DELETE FROM transactions WHERE id = ?", (tid,))
                await db.commit()
            await callback.message.answer("✅ Транзакция удалена", reply_markup=get_main_menu_buttons())

    except Exception as e:
        logger.error(f"Ошибка в handle_callback: {e}")
        try:
            await callback.answer("Произошла ошибка, попробуйте позже")
        except:
            pass


# --- Обработка ввода категории ---
@dp.message(TransactionStates.waiting_for_category)
async def process_category(message: types.Message, state: FSMContext):
    category = message.text.strip()
    if len(category) > 30:
        await message.answer("Категория слишком длинная. Максимум 30 символов.")
        return

    data = await state.get_data()
    if 'filter_type' in data:  # Если это фильтр по категории для отчета
        await state.clear()
        await send_report(message, data['report_period'], 'category', category)
    else:  # Если это добавление транзакции
        await state.update_data(category=category)
        await state.set_state(TransactionStates.waiting_for_amount)
        t_type = data['t_type']

        # Сохраняем категорию для быстрого доступа
        async with aiosqlite.connect("finance.db") as db:
            await db.execute('''
                INSERT OR IGNORE INTO categories (user_id, category_type, category_name)
                VALUES (?, ?, ?)
            ''', (message.from_user.id, t_type, category))
            await db.commit()

        await message.answer(
            f"Введите сумму {t_type}а (только число, например: 1500):",
            reply_markup=get_cancel_button()
        )


# --- Обработка ввода суммы ---
@dp.message(TransactionStates.waiting_for_amount)
async def process_amount(message: types.Message, state: FSMContext):
    try:
        amount = float(message.text.strip().replace(",", "."))
        if amount <= 0:
            await message.answer("Сумма должна быть больше нуля.")
            return

        data = await state.get_data()
        t_type = data['t_type']
        category = data['category']

        async with aiosqlite.connect("finance.db") as db:
            await db.execute('''
                INSERT INTO transactions (user_id, type, amount, category, description)
                VALUES (?, ?, ?, ?, ?)
            ''', (message.from_user.id, t_type, amount, category, ""))
            await db.commit()

        await state.clear()
        await message.answer(
            f"{'✅ Доход' if t_type == 'доход' else '💸 Расход'} записан:\n"
            f"Категория: {category}\n"
            f"Сумма: {format_number(amount)}₽",
            reply_markup=get_main_menu_buttons()
        )
    except ValueError:
        await message.answer("Неверный формат суммы. Введите число (например: 1500 или 99.90).")

# --- Обработка ввода даты для отчета ---
@dp.message(TransactionStates.waiting_for_start_date)
async def process_start_date(message: types.Message, state: FSMContext):
    try:
        date_str = message.text.strip()
        datetime.strptime(date_str, "%Y-%m-%d")  # Проверка формата
        await state.update_data(start_date=date_str)

        data = await state.get_data()
        if data.get('excel_report'):
            await state.set_state(TransactionStates.waiting_for_end_date)
            await message.answer(
                "Введите конечную дату в формате ГГГГ-ММ-ДД (или 'нет' для отчета за один день):",
                reply_markup=get_cancel_button()
            )
        else:
            await state.set_state(TransactionStates.waiting_for_end_date)
            await message.answer(
                "Введите конечную дату в формате ГГГГ-ММ-ДД:",
                reply_markup=get_cancel_button()
            )
    except ValueError:
        await message.answer("Неверный формат даты. Введите в формате ГГГГ-ММ-ДД (например: 2023-12-31).")


@dp.message(TransactionStates.waiting_for_end_date)
async def process_end_date(message: types.Message, state: FSMContext):
    try:
        end_date_str = message.text.strip().lower()
        data = await state.get_data()

        if data.get('excel_report') and end_date_str == 'нет':
            end_date_str = data['start_date']
        else:
            datetime.strptime(end_date_str, "%Y-%m-%d")  # Проверка формата

        start_date = data['start_date']

        if end_date_str < start_date:
            await message.answer("Конечная дата должна быть после начальной.")
            return

        if data.get('excel_report'):
            await state.clear()
            await generate_excel_report(
                message.from_user.id,
                message,
                start_date,
                end_date_str
            )
        else:
            await state.clear()
            await send_report(
                message,
                "custom",
                "all",
                custom_start=start_date,
                custom_end=end_date_str
            )
    except ValueError:
        await message.answer("Неверный формат даты. Введите в формате ГГГГ-ММ-ДД (например: 2023-12-31).")


# --- Генерация Excel отчета ---
async def generate_excel_report(user_id, message, start_date, end_date):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        # Заголовки
        headers = ["Дата", "Тип", "Категория", "Сумма", "Описание"]
        ws.append(headers)

        # Получаем данные из БД
        async with aiosqlite.connect("finance.db") as db:
            cursor = await db.execute("""
                SELECT 
                    timestamp, 
                    type, 
                    category, 
                    amount, 
                    COALESCE(description, '') as description
                FROM transactions
                WHERE user_id = ? AND DATE(timestamp) BETWEEN ? AND ?
                ORDER BY timestamp
            """, (user_id, start_date, end_date))
            rows = await cursor.fetchall()

        if not rows:
            await message.answer(
                "📭 Нет данных за выбранный период",
                reply_markup=get_main_menu_buttons()
            )
            return

        # Заполняем данные
        total_income = 0
        total_expense = 0
        income_categories = {}
        expense_categories = {}

        for row in rows:
            timestamp, t_type, category, amount, description = row
            date_str = timestamp.split()[0]

            ws.append([
                date_str,
                t_type.capitalize(),
                category,
                amount,
                description
            ])

            if t_type == "доход":
                total_income += amount
                if category not in income_categories:
                    income_categories[category] = 0
                income_categories[category] += amount
            else:
                total_expense += amount
                if category not in expense_categories:
                    expense_categories[category] = 0
                expense_categories[category] += amount

        # Добавляем итоги
        ws.append([])
        ws.append(["Итого:"])
        ws.append(["Доходы:", total_income])
        ws.append(["Расходы:", total_expense])
        ws.append(["Баланс:", total_income - total_expense])

        # Создаем лист для статистики
        stats_ws = wb.create_sheet(title="Статистика")
        stats_ws.append(["Категория", "Тип", "Сумма"])

        # Подготавливаем данные для диаграмм
        chart_row = 1
        for category, amount in income_categories.items():
            stats_ws.append([category, "Доход", amount])
            chart_row += 1

        for category, amount in expense_categories.items():
            stats_ws.append([category, "Расход", amount])
            chart_row += 1

        # Диаграмма распределения доходов (круговая)
        if income_categories:
            income_pie = openpyxl.chart.PieChart()
            income_labels = openpyxl.chart.Reference(stats_ws,
                                                     min_col=1,
                                                     min_row=2,
                                                     max_row=len(income_categories) + 1)
            income_data = openpyxl.chart.Reference(stats_ws,
                                                   min_col=3,
                                                   min_row=1,
                                                   max_row=len(income_categories) + 1)
            income_pie.title = "Распределение доходов по категориям"
            income_pie.add_data(income_data, titles_from_data=True)
            income_pie.set_categories(income_labels)
            income_pie.width = 15
            income_pie.height = 9
            ws.add_chart(income_pie, "G2")

        # Диаграмма распределения расходов (круговая)
        if expense_categories:
            expense_pie = openpyxl.chart.PieChart()
            expense_labels = openpyxl.chart.Reference(stats_ws,
                                                      min_col=1,
                                                      min_row=len(income_categories) + 2,
                                                      max_row=chart_row)
            expense_data = openpyxl.chart.Reference(stats_ws,
                                                    min_col=3,
                                                    min_row=len(income_categories) + 1,
                                                    max_row=chart_row)
            expense_pie.title = "Распределение расходов по категориям"
            expense_pie.add_data(expense_data, titles_from_data=True)
            expense_pie.set_categories(expense_labels)
            expense_pie.width = 15
            expense_pie.height = 9
            ws.add_chart(expense_pie, "G20")

        # Гистограмма сравнения доходов/расходов
        if income_categories and expense_categories:
            bar_chart = openpyxl.chart.BarChart()
            bar_chart.title = "Сравнение доходов и расходов"
            bar_chart.style = 10
            bar_chart.y_axis.title = "Сумма"
            bar_chart.x_axis.title = "Категории"
            bar_chart.width = 25
            bar_chart.height = 12

            data = openpyxl.chart.Reference(stats_ws,
                                            min_col=3,
                                            min_row=1,
                                            max_row=chart_row)
            cats = openpyxl.chart.Reference(stats_ws,
                                            min_col=1,
                                            min_row=2,
                                            max_row=chart_row)

            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            ws.add_chart(bar_chart, "G38")

        # Форматирование
        for sheet in wb:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Отправляем файл
        await message.answer_document(
            types.BufferedInputFile(
                buffer.read(),
                filename=f"finance_report_{start_date}_{end_date}.xlsx"
            ),
            caption=f"📊 Полный финансовый отчет\n"
                    f"📅 Период: {start_date} - {end_date}\n"
                    f"🟢 Доходы: {format_number(total_income)}₽ ({len(income_categories)} категорий)\n"
                    f"🔴 Расходы: {format_number(total_expense)}₽ ({len(expense_categories)} категорий)\n"
                    f"💰 Итоговый баланс: {format_number(total_income - total_expense)}₽",
            reply_markup=get_main_menu_buttons()
        )
        buffer.close()

    except Exception as e:
        logger.error(f"Ошибка при генерации Excel отчета: {e}")
        await message.answer(
            "Произошла ошибка при генерации отчета",
            reply_markup=get_main_menu_buttons()
        )


# --- Получение и отправка отчета ---
async def send_report(source, period: str, filter_type: str, category_filter=None, custom_start=None, custom_end=None):
    if isinstance(source, types.CallbackQuery):
        user_id = source.from_user.id
        message = source.message
    else:
        user_id = source.from_user.id
        message = source

    now = datetime.now()

    if period == "day":
        where = "DATE(timestamp)=?"
        params = [now.strftime("%Y-%m-%d")]
        title = "📅 Сегодня"
    elif period == "month":
        where = "strftime('%Y-%m', timestamp)=?"
        params = [now.strftime("%Y-%m")]
        title = "🗓 Этот месяц"
    elif period == "quarter":
        quarter = (now.month - 1) // 3 + 1
        months = [(quarter - 1) * 3 + i for i in range(1, 4)]
        placeholders = ','.join('?' * len(months))
        where = f"strftime('%Y', timestamp)=? AND cast(strftime('%m', timestamp) as int) IN ({placeholders})"
        params = [str(now.year)] + months
        title = f"📊 {quarter}-й квартал"
    elif period == "year":
        where = "strftime('%Y', timestamp)=?"
        params = [str(now.year)]
        title = "📆 Этот год"
    elif period == "custom":
        where = "DATE(timestamp) BETWEEN ? AND ?"
        params = [custom_start, custom_end]
        title = f"📅 Период с {custom_start} по {custom_end}"
    else:
        await message.answer("❌ Неверный период")
        return

    query = f'''
        SELECT type, category, amount, timestamp
        FROM transactions
        WHERE user_id=? AND {where}
    '''
    if filter_type == "income":
        query += " AND type='доход'"
        title += " – Доходы"
    elif filter_type == "expense":
        query += " AND type='расход'"
        title += " – Расходы"
    elif filter_type == "category" and category_filter:
        query += " AND category=?"
        params.append(category_filter)
        title += f" – Категория: {category_filter}"
    else:
        title += " – Все транзакции"

    async with aiosqlite.connect("finance.db") as db:
        cursor = await db.execute(query, (user_id, *params))
        rows = await cursor.fetchall()

    if not rows:
        await message.answer(
            f"{title}\n\n📭 Нет данных за выбранный период",
            reply_markup=get_main_menu_buttons()
        )
        return

    total_income = 0
    total_expense = 0
    categories = {}
    text = f"<b>{title}</b>\n\n"

    for t_type, category, amount, ts in rows:
        emoji = "🟢" if t_type == "доход" else "🔴"
        text += f"{emoji} <b>{t_type.capitalize()}</b> – {category}: {format_number(amount)}₽\n🕒 {ts}\n\n"

        if t_type == "доход":
            total_income += amount
        else:
            total_expense += amount

        # Статистика по категориям
        if category not in categories:
            categories[category] = {"income": 0, "expense": 0}
        if t_type == "доход":
            categories[category]["income"] += amount
        else:
            categories[category]["expense"] += amount

    balance = total_income - total_expense
    text += "<b>Итоги:</b>\n"
    text += f"🟢 Доходов: {format_number(total_income)}₽\n"
    text += f"🔴 Расходов: {format_number(total_expense)}₽\n"
    text += f"💰 Баланс: {format_number(balance)}₽ {'(плюс)' if balance >= 0 else '(минус)'}\n\n"

    # Добавляем статистику по категориям
    if filter_type != "category":
        text += "<b>По категориям:</b>\n"
        for category, amounts in categories.items():
            cat_balance = amounts["income"] - amounts["expense"]
            if amounts["income"] > 0 and amounts["expense"] > 0:
                text += f"🏷 {category}: +{format_number(amounts['income'])} / -{format_number(amounts['expense'])} = {format_number(cat_balance)}₽\n"
            elif amounts["income"] > 0:
                text += f"🏷 {category}: +{format_number(amounts['income'])}₽\n"
            else:
                text += f"🏷 {category}: -{format_number(amounts['expense'])}₽\n"

    await message.answer(text, parse_mode="HTML", reply_markup=get_main_menu_buttons())


# --- Запуск ---
async def main():
    await init_db()
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())